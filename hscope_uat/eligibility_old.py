from langchain_ollama import ChatOllama
from langchain.prompts import ChatPromptTemplate
import pymupdf4llm
import re
from collections import OrderedDict
import json
import os
import pandas as pd
from typing import Dict, Any, List
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from hscope_uat.UAT_config import*

# ============================================================
# Initialize Chat
# ============================================================
chat = ChatOllama(
    model="gpt-oss:20b",
    temperature=0.0,
    top_p=1.0,
    max_tokens=1024,
    verbose=False
)

# ============================================================
# Hardcoded Age Dictionary
# ============================================================
AGE_DICT = {
    "Employee": {"Min_Age_Years": 18, "Min_Age_Months": None, "Max_Age_Years": 65, "Max_Age_Months": None, "Type": "adult"},
    "Spouse": {"Min_Age_Years": 18, "Min_Age_Months": None, "Max_Age_Years": 65, "Max_Age_Months": None, "Type": "adult"},
    "Father": {"Min_Age_Years": 40, "Min_Age_Months": None, "Max_Age_Years": 100, "Max_Age_Months": None, "Type": "adult"},
    "Mother": {"Min_Age_Years": 40, "Min_Age_Months": None, "Max_Age_Years": 100, "Max_Age_Months": None, "Type": "adult"},
    "Father inlaw": {"Min_Age_Years": 40, "Min_Age_Months": None, "Max_Age_Years": 100, "Max_Age_Months": None, "Type": "adult"},
    "Mother inlaw": {"Min_Age_Years": 40, "Min_Age_Months": None, "Max_Age_Years": 100, "Max_Age_Months": None, "Type": "adult"},
    "Son": {"Min_Age_Years": 0, "Min_Age_Months": 3, "Max_Age_Years": 25, "Max_Age_Months": None, "Type": "child"},
    "Daughter": {"Min_Age_Years": 0, "Min_Age_Months": 3, "Max_Age_Years": 25, "Max_Age_Months": None, "Type": "child"},
}

# ============================================================
# PDF → Markdown
# ============================================================
def pdf_to_md(pdf_path):
    return pymupdf4llm.to_markdown(pdf_path)

# ============================================================
# Normalize ID
# ============================================================
def normalize_id(eid: str) -> str:
    m = re.match(r"^(\d+)(?:\(?([a-zivx]+)\)?)?$", eid, re.IGNORECASE)
    if not m:
        return eid
    num, suffix = m.groups()
    if suffix:
        return f"{num}({suffix.lower()})"
    return num

# ============================================================
# Extract endorsements from Markdown
# ============================================================
def extract_endorsements(md_text: str, debug: bool = False):
    lines = md_text.splitlines()
    heading_pattern = re.compile(
        r"^\s*\*{0,2}\s*(?:Endorsement|Endt\.?)\.?\s*No\.?\s*"
        r"(\d+)"
        r"(?:\s*\(\s*([A-Za-z0-9ivxIVX]{1,3})\s*\)"
        r"|-(?:([A-Za-z0-9ivxIVX]{1,3}))"
        r"|([A-Za-z]{1,3})(?=\s|[-–]|$))?",
        re.IGNORECASE,
    )

    junk_patterns = [
    re.compile(r"^group health policy", re.IGNORECASE),
    re.compile(r"^uin[:\s]", re.IGNORECASE),
    re.compile(r"^irda", re.IGNORECASE),
    re.compile(r"^policy number", re.IGNORECASE),
    re.compile(r"^name of the insured", re.IGNORECASE),
    re.compile(r"^period of insurance", re.IGNORECASE),
    re.compile(r"endorsements attached", re.IGNORECASE),
    re.compile(r"Group Health Policy – Endorsements", re.IGNORECASE),
    re.compile(r"^Page\s+\*\*\d+\*\*\s+of\s+\*\*\d+\*\*$", re.IGNORECASE),
    re.compile(r"^\*\*Policy Number.*\*\*$", re.IGNORECASE),
    re.compile(r"^\*\*Name of the Insured.*\*\*$", re.IGNORECASE),
    re.compile(r"^\*\*Period of Insurance.*\*\*$", re.IGNORECASE),
    re.compile(r"^//\s*\d+\s*//$", re.IGNORECASE),
    re.compile(r"^royal sundaram.*", re.IGNORECASE),
    re.compile(r"^regd office.*", re.IGNORECASE),
    re.compile(r"^corporate office.*", re.IGNORECASE),
    re.compile(r"^email[:\s].*", re.IGNORECASE),
    re.compile(r"^website[:\s].*", re.IGNORECASE),
    re.compile(r"^ph[:\s].*", re.IGNORECASE),
    re.compile(r"^.*irda regn.*", re.IGNORECASE),
    re.compile(r"^.*cin[-:\s].*", re.IGNORECASE),
    re.compile(r".*\bchennai\s*\d{3}\s*\d{3}.*", re.IGNORECASE)
]


    endorsements = OrderedDict()
    current_id, buffer = None, []

    def clean_buffer(buf):
        return [line for line in buf if not any(p.search(line) for p in junk_patterns)]
    for line in lines:
        m = heading_pattern.search(line)
        if m:
            num = m.group(1)
            suffix = m.group(2) or m.group(3) or m.group(4)
            canonical_id = num
            if suffix and len(suffix) <= 3:
                canonical_id = f"{num}({suffix.lower()})"
            if current_id and buffer:
                endorsements[current_id] = "\n".join(clean_buffer(buffer)).strip()
            current_id = canonical_id
            buffer = [line]
        elif current_id:
            buffer.append(line)

    if current_id and buffer:
        endorsements[current_id] = "\n".join(clean_buffer(buffer)).strip()
    return endorsements

# ============================================================
# Extract Special Conditions
# ============================================================
def extract_special_conditions(md_text: str) -> Dict[str, List[str]]:
    """
    Extract blocks starting from bold markdown headers that contain keywords:
    'special' or 'other' + ('condition'/'conditions'/'clause'/'clauses'/'coverage').
    Captures from that header until the next bold header.
    """
    header_pattern = re.compile(r"\*\*(.+?)\*\*", re.IGNORECASE)
    matches = list(header_pattern.finditer(md_text))
    results = []

    for i, match in enumerate(matches):
        header_text = match.group(1).strip().lower()
        if (("special" in header_text or "other" in header_text) and
            any(word in header_text for word in ["condition", "conditions", "clause", "clauses", "coverage","Endorsements"," Endorsement","Note","Notes"])):
            start = match.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(md_text)
            block = md_text[start:end].strip()
            block_cleaned = re.sub(r"\*\*(.+?)\*\*", r"\1", block)
            results.append(block_cleaned)

    endorsements_context = ""
    for endnt in results:
        endorsements_context += endnt + "\n\n"
    return {"special condition": endorsements_context}


def overaly_specail_condtion(md_text):
    endnt_context = extract_special_conditions(md_text=md_text)
    if endnt_context['special condition'] == "":
        spl_cond_context = """There is no special condition provided, so you can consider No context from special conditions section."""
        endnt_context['special condition'] = spl_cond_context
        return endnt_context
    return endnt_context

# ============================================================
# Get endorsements content
# ============================================================
def get_endorsements_content(pdf_path, desired_endorsements):
    md_text = pdf_to_md(pdf_path)
    endorsements = extract_endorsements(md_text)
    desired_endorsements = [normalize_id(eid) for eid in desired_endorsements]
    matching_endorsements = {
        eid: endorsements[eid]
        for eid in desired_endorsements
        if eid in endorsements
    }

    # Also extract special conditions here
    special_conditions_context = overaly_specail_condtion(md_text)
    return list(matching_endorsements.values()), special_conditions_context
# ============================================================
# Relationship Extractor (LLM)
# ============================================================
def relationship_extractor(sample_text: str, chat, special_conditions: str) -> Dict[str, Any]:
    combined_context = f"{sample_text}\n\n---\n\nAdditional Special Conditions Context:\n{special_conditions}"
    
    prompt = ChatPromptTemplate.from_messages([
        ("system",
         "You are an expert insurance policy analyzer. "
         "Extract family relationships from endorsement clauses and return structured JSON only. "
         "Be precise and follow the rules exactly."
        ),
        ("human",
         """
         Analyze the following endorsement text and its special condition context:

         {context}

         Return strictly valid JSON in this exact format:
         {{
           "relationships": ["Employee", "Spouse", "Son", "Daughter", "Father", "Mother", "Father inlaw", "Mother inlaw"],
           "max_members": 8
         }}

         Special Rules (Follow these rules strictly):
         1. Always include "Employee" as the first relationship.
         
         2. If spouse is mentioned, include "Spouse".
         
         3. For dependent children - IMPORTANT DEFAULT RULE:
            - ALWAYS use exactly 2 children: "Son" and "Daughter" as the DEFAULT.
            - This applies regardless of what maximum number is stated in the policy (even if it says 3, 4, or any other number).
            - Do NOT add more than 2 children under any circumstances.
            - If children are mentioned, always include both "Son" and "Daughter".
         
         4. For parents - THIS IS CRITICAL:
            - If text mentions ONLY "Dependent Parents" (without any mention of "in law", "in-law", or "inlaw"), add ONLY "Father" and "Mother" as separate entries.
            - If text mentions "Dependent Parents in law", "Parents in law", "Dependant Parents in law", or uses "/" or "and" connecting parents with in-laws (e.g., "Dependent Parents/Dependant Parents in law" or "Dependent Parents and Dependent Parents in law"), add ALL FOUR: "Father", "Mother", "Father inlaw", and "Mother inlaw".
            - Each parent is counted as a separate member.
         
         5. The max_members must equal the total count of ALL relationships in the array.
         
         6. Do not add any relationships not explicitly mentioned in the policy (except use the 2-children default rule when children are mentioned).
         
         7. Use exact relationship names: "Father inlaw" and "Mother inlaw" (not "Father-in-law" or "Father in-law").

         8. Summary of typical outputs:
            - Employee = ["Employee"] = 1 Members
            - Employee + Spouse = ["Employee","Spouse"] = 2 Members
            - Employee + Spouse + Children only = ["Employee", "Spouse", "Son", "Daughter"] = 4 members
            - Employee + Spouse + Children + Parents = ["Employee", "Spouse", "Son", "Daughter", "Father", "Mother"] = 6 members
            - Employee + Spouse + Children + Parents + In-laws = ["Employee", "Spouse", "Son", "Daughter", "Father", "Mother", "Father inlaw", "Mother inlaw"] = 8 members
         
         Return ONLY the JSON object, no explanation or additional text.
         """)
    ])
    
    chain = prompt | chat
    
    try:
        result = chain.invoke({"context": combined_context})
        text = (result.content or "").strip()
        
        # Remove markdown code blocks if present
        if text.startswith("```"):
            text = re.sub(r"^```(?:json|JSON)?\s*", "", text, flags=re.IGNORECASE)
            text = re.sub(r"\s*```\s*$", "", text)
            text = text.strip()
        
        # Find the JSON object
        first = text.find("{")
        last = text.rfind("}")
        
        if first == -1 or last == -1:
            raise ValueError("No JSON object found in response")
        
        json_text = text[first:last + 1]
        parsed_data = json.loads(json_text)
        
        # Validate the structure
        if "relationships" not in parsed_data or "max_members" not in parsed_data:
            raise ValueError("Missing required fields in JSON")
        
        # Ensure max_members is an integer
        if isinstance(parsed_data["max_members"], str):
            parsed_data["max_members"] = int(parsed_data["max_members"])
        
        # Auto-correct max_members to match relationship count
        actual_count = len(parsed_data["relationships"])
        if parsed_data["max_members"] != actual_count:
            print(f"Warning: max_members ({parsed_data['max_members']}) doesn't match relationship count ({actual_count}). Auto-correcting to {actual_count}.")
            parsed_data["max_members"] = actual_count
        
        return parsed_data
        
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
        print(f"Attempted to parse: {json_text if 'json_text' in locals() else text}")
        return {"relationships": ["Employee"], "max_members": 1}
    except Exception as e:
        print(f"Error in relationship extraction: {e}")
        return {"relationships": ["Employee"], "max_members": 1}

# ============================================================
# Build relationship schema
# ============================================================
def build_relationship_schema(extracted: Dict[str, Any], endorsement_text: str, chat) -> Dict[str, Any]:
    relationships = extracted.get("relationships", [])
    relationship_list = []
    min_ages_years, min_ages_months = [], []
    max_ages_years, max_ages_months = [], []
    member_count, member_type = [], []

    for rel in relationships:
        rel_clean = rel.strip().lower()
        if rel_clean == "children":
            for child in ["Son", "Daughter"]:
                relationship_list.append(child)
                age_info = AGE_DICT[child]
                min_ages_years.append(age_info["Min_Age_Years"])
                min_ages_months.append(age_info["Min_Age_Months"])
                max_ages_years.append(age_info["Max_Age_Years"])
                max_ages_months.append(age_info["Max_Age_Months"])
                member_count.append(1)
                member_type.append(age_info["Type"])
        elif "in law" in rel_clean:
            for parent in ["Father", "Mother", "Father inlaw", "Mother inlaw"]:
                relationship_list.append(parent)
                age_info = AGE_DICT[parent]
                min_ages_years.append(age_info["Min_Age_Years"])
                min_ages_months.append(age_info["Min_Age_Months"])
                max_ages_years.append(age_info["Max_Age_Years"])
                max_ages_months.append(age_info["Max_Age_Months"])
                member_count.append(1)
                member_type.append(age_info["Type"])
        elif "parent" in rel_clean:
            for parent in ["Father", "Mother"]:
                relationship_list.append(parent)
                age_info = AGE_DICT[parent]
                min_ages_years.append(age_info["Min_Age_Years"])
                min_ages_months.append(age_info["Min_Age_Months"])
                max_ages_years.append(age_info["Max_Age_Years"])
                max_ages_months.append(age_info["Max_Age_Months"])
                member_count.append(1)
                member_type.append(age_info["Type"])
        else:
            relationship_list.append(rel)
            age_info = AGE_DICT.get(rel, {})
            min_ages_years.append(age_info.get("Min_Age_Years"))
            min_ages_months.append(age_info.get("Min_Age_Months"))
            max_ages_years.append(age_info.get("Max_Age_Years"))
            max_ages_months.append(age_info.get("Max_Age_Months"))
            member_count.append(1)
            member_type.append(age_info.get("Type"))

    return {
        "Max No Of Members Covered": len(relationship_list),
        "Relationship Covered in Text": relationship_list,
        "Min_Age(InYears)": min_ages_years,
        "Min_Age(In Months)": min_ages_months,
        "Max_Age(InYears)": max_ages_years,
        "Max_Age(In Months)": max_ages_months,
        "Member_Count": member_count,
        "Member_Type": member_type,
        "Sublimit_Applicable": "No",
        "Sublimit_Type": "",
        "Sub_Limit": "",
        "Family Buffer Applicable": "No",
        "Family Buffer Amount": "",
        "Is Network Applicable": "No",
        "Black listed hospitals are applicable?": "Yes"
    }

# ============================================================
# Corporate Buffer + CI Extractor
# ============================================================
def buffer_and_ci_extractor(sample_text: str, chat, special_conditions: str) -> Dict[str, Any]:
    combined_context = f"{sample_text}\n\n---\n\nAdditional Special Conditions Context:\n{special_conditions}"
    prompt = ChatPromptTemplate.from_messages([
        ("system",
         "You are an expert insurance policy analyzer. "
         "Extract Corporate Buffer Eligibility and Critical Illness details. "
         "Follow the exact rules for Corporate Buffer and CI eligibility mapping."
        ),
        ("human",
         """
         Analyze the following endorsement text:

         {context}

         Return strictly valid JSON in this exact format:
         {{
           "Corporate Buffer Eligibility": {{
             "Corporate Buffer applicable": "Yes/No",
             "Buffer Type": "If  Corporate Buffer Eligibility field is "Yes" then this field is mentioned as "Both" else ' ' "
             "Applicable for": "If  Corporate Buffer Eligibility field is "Yes" then this field is mentioned as "Both" else ' ' "
             "Total Corporate Buffer": "Fetch the Corporate Buffer amount which is next to the keyword 'limit of Rs'"
             "Corporate Buffer Limit Per Family": "", return this as null by default
             "Corporate Buffer Limit Per Parent": "", return this as null by default
             "Reload of SI": "Mapping rules for Reload of SI(Follow this rules strictly)"
             "Approving Authority_1":"If  Corporate Buffer Eligibility field is "Yes" then this field is mentioned as "Corporate HR" else ' ' "
             "Buffer OPD Limit": "",return this as null by default
             "Whether increase in sum insured permissible at renewal_CB": "If  Corporate Buffer Eligibility field is "Yes" then this field is mentioned as "NO" else ' ' "
           }},
           "Critical Illness Eligible": {{
             "Critical Illness applicable": "In the extracted special condition context check whether any context with 'Critical Illness' And amount or percentage is mentioned — if YES then 'Yes' else 'No'.",
             "Critical Illness limit per family": "If present—Extract numeric amount or percentage mentioned, else ' '.",
             "Approving Authority_2": "If  Critical Illness Eligible field is "Yes" then this field is mentioned as "Corporate HR" else ' ' "
             "Whether increase in sum insured permissible at renewal_CL": "If  Critical Illness Eligible field is "Yes" then this field is mentioned as "NO" else ' ' "
           }}
         }}

         Mapping rules for Reload of SI(Follow this rules strictly):
         - if text says "equivalent to the per person limit / per Insured person limit" → "Reload of SI is up to the Existing SI"
         - if text says "equivalent to twice/double" → "Reload of SI is up to Double the SI"
         - if text says "equivalent to thrice" → "Reload of SI is up to the Thrice the SI"
         - if text says "No limit" → "No limit for the reload of SI"

         Note:
         - In corporate buffer amount fields, only numeric amounts should appear — do not include 'Rs.', 'Rupees', or any currency text.
         """)
    ])
    chain = prompt | chat
    result = chain.invoke({"context": combined_context})
    text = (result.content or "").strip()
    try:
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text).rstrip("`\n\r ")
            text = re.sub(r"\s*```$", "", text).strip()
        first, last = text.find("{"), text.rfind("}")
        json_text = text[first:last + 1] if first != -1 and last != -1 else text
        return json.loads(json_text)
    except Exception:
                return {
            "Corporate Buffer Eligibility": {
                "Corporate Buffer applicable": "No",
                "Buffer Type": "",
                "Applicable for": "",
                "Total Corporate Buffer": "",
                "Corporate Buffer Limit Per Family": "",
                "Corporate Buffer Limit Per Parent": "",
                "Reload of SI": "",
                "Approving Authority_1":"",#
                "Buffer OPD Limit": "",
                "Whether increase in sum insured permissible at renewal_CB": ""
            },
            "Critical Illness Eligible": {
                "Critical Illness applicable": "",
                "Critical Illness limit per family": "",
                "Approving Authority_2": "",
                "Whether increase in sum insured permissible at renewal_CL": ""
            }
        }


# ============================================================
# Save to Excel (expanded + remove duplicates)
# ============================================================
def format_for_excel(value):
    if isinstance(value, list):
        return "\n".join(str(v) for v in value if v not in (None, "None", "", " ")).strip()
    if value in (None, "None"):
        return ""
    return str(value).strip()

def save_combined_to_excel(rel_json: dict, buffer_json: dict, pdf_path: str,source_dir):
    combined = {**rel_json, **buffer_json.get("Corporate Buffer Eligibility", {}), **buffer_json.get("Critical Illness Eligible", {})}

    list_lengths = [len(v) for v in combined.values() if isinstance(v, list)]
    max_len = max(list_lengths) if list_lengths else 1

    single_occurrence_fields = [
        "Max No Of Members Covered",
        "Family Buffer Applicable",
        "Is Network Applicable",
        "Black listed hospitals are applicable?",
        "Corporate Buffer applicable",
        "Buffer Type",
        "Applicable for",
        "Total Corporate Buffer",
        "Reload of SI",
        "Approving Authority_1",#
        "Critical Illness applicable",
        "Whether increase in sum insured permissible at renewal_CB",
        "Whether increase in sum insured permissible at renewal_CL",
        "Approving Authority_2",
        "Critical Illness limit per family"

    ]



    rows = []
    for i in range(max_len):
        row = {}
        for k, v in combined.items():
            if isinstance(v, list):
                cell_val = v[i] if i < len(v) else ""
                row[k] = format_for_excel(cell_val)
            else:
                if k in single_occurrence_fields and i > 0:
                    row[k] = ""
                else:
                    row[k] = format_for_excel(v)
        rows.append(row)

    df = pd.DataFrame(rows)
    
    output_folder = source_dir  # define your working folder
    os.makedirs(output_folder, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    excel_file = os.path.join(os.path.dirname(pdf_path), f"{base_name}_eligibility_sheet.xlsx")
    df.to_excel(excel_file, index=False)

    wb = load_workbook(excel_file)
    ws = wb.active
    for i, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            if cell.value:
                for line in str(cell.value).splitlines():
                    if len(line) > max_len:
                        max_len = len(line)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 100)
    wb.save(excel_file)
    print(f"\n✅ Final Excel saved at: {excel_file}")
    return excel_file

# # ============================================================
# # Main Driver
# # ============================================================
# def main_driver(pdf_path: str, chat):
#     endt1_texts, special_conditions = get_endorsements_content(pdf_path, ["1", "1a", "1b", "1c", "1d"])
#     if not endt1_texts:
#         print("No 1-series endorsements found.")
#         return
#     rel_extracted = relationship_extractor(endt1_texts[0], chat, special_conditions['special condition'])
#     rel_json = build_relationship_schema(rel_extracted, endt1_texts[0], chat)

#     endt10_texts, special_conditions_10 = get_endorsements_content(pdf_path, ["10", "10a", "10b", "10c", "10d"])
#     buffer_json = buffer_and_ci_extractor(endt10_texts[0], chat, special_conditions_10['special condition']) if endt10_texts else {
#         "Corporate Buffer Eligibility": {
#             "Corporate Buffer applicable": "No",
#             "Buffer Type": "",
#             "Applicable for": "",
#             "Total Corporate Buffer": "",
#             "Corporate Buffer Limit Per Family": "",
#             "Corporate Buffer Limit Per Parent": "",
#             "Reload of SI": "",
#             "Buffer OPD Limit": "",
#             "Whether increase in sum insured permissible at renewal_CB": ""
#         },
#         "Critical Illness Eligible": {
#             "Critical Illness applicable": "No",
#             "Critical Illness limit per family": "",
#             "Approving Authority": "",
#             "Whether increase in sum insured permissible at renewal_CL": ""
#         }
# }

#     save_combined_to_excel(rel_json, buffer_json, pdf_path)

def generate_eligibility(pdf_path: str, chat,source_dir):
    # ============================================================
    # 1️⃣ Extract Endorsement 1-series (Relationship Info)
    # ============================================================
    endt1_texts, special_conditions = get_endorsements_content(pdf_path, ["1", "1a", "1b", "1c", "1d"])
    if not endt1_texts:
        print("No 1-series endorsements found.")
        return

    rel_extracted = relationship_extractor(
        endt1_texts[0],
        chat,
        special_conditions['special condition']
    )
    rel_json = build_relationship_schema(rel_extracted, endt1_texts[0], chat)

    # ============================================================
    # 2️⃣ Extract Endorsement 10-series (Buffer + CI Info)
    # ============================================================
    endt10_texts, special_conditions_10 = get_endorsements_content(pdf_path, ["10", "10a", "10b", "10c", "10d"])

    if endt10_texts:
        # ✅ Normal case: 10-series endorsement is present
        buffer_json = buffer_and_ci_extractor(
            endt10_texts[0],
            chat,
            special_conditions_10['special condition']
        )

    else:
        # ⚠️ 10-series not found — check if Special Condition still has CI info
        special_condition_text = special_conditions_10.get('special condition', '').strip()

        if special_condition_text and "critical illness" in special_condition_text.lower():
            print("🔍 No Endt. 10 found, but Special Condition mentions Critical Illness — analyzing...")
            buffer_json = buffer_and_ci_extractor(
                special_condition_text,
                chat,
                special_condition_text
            )
        else:
            print("⚠️ No 10-series endorsement or relevant Critical Illness context found.")
            buffer_json = {
                "Corporate Buffer Eligibility": {
                    "Corporate Buffer applicable": "No",
                    "Buffer Type": "",
                    "Applicable for": "",
                    "Total Corporate Buffer": "",
                    "Corporate Buffer Limit Per Family": "",
                    "Corporate Buffer Limit Per Parent": "",
                    "Reload of SI": "",
                    "Approving Authority_1": "",#
                    "Buffer OPD Limit": "",
                    "Whether increase in sum insured permissible at renewal_CB": ""
                },
                "Critical Illness Eligible": {
                    "Critical Illness applicable": "No",
                    "Critical Illness limit per family": "",
                    "Approving Authority_2": "",
                    "Whether increase in sum insured permissible at renewal_CL": ""
                }
            }


    # ============================================================
    # 3️⃣ Save Combined Excel Output
    # ============================================================
    save_combined_to_excel(rel_json, buffer_json, pdf_path,source_dir)

# ============================================================
# Run
# ============================================================
if __name__ == "__main__":
    pdf_path = "/home/ubuntu/rspdf/elgibility_sheet/Dinesh/input/1a_input/HG00006790000100.pdf"
    source_dir=r""
    generate_eligibility(pdf_path, chat,source_dir)